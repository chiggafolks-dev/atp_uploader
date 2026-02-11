
# atp_text_insert.py (updated)
import openpyxl
import re

class ATPTextReplacer:
    def __init__(self, workbook):
        self.wb = workbook
        
        # NEW: Use regex patterns to match various placeholder formats
        self.TEXT_PLACEHOLDER_PATTERNS = {
            r'\[SITE_?ID\]': 'site_id',          # Matches [SITE_ID] or [SITEID]
            r'\[SITE_?NAME\]': 'site_name',      # Matches [SITE_NAME] or [SITENAME]
            r'\[HOSTNAME\]': 'hostname',
            r'\[SCOPE_?OF_?WORK\]': 'scope_of_work',
            r'\[DEVICE_?TYPE\]': 'device_type',
            r'\[PROJECT_?CODE\]': 'project_code',
            r'\[DATE\]': 'date',
            r'\[ENGINEER\]': 'engineer',
            # Add more patterns as needed
        }

    def replace(self, values: dict):
        """
        Replace all text placeholders in the workbook with actual values.
        Now supports various placeholder formats using regex.
        """
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell_text = str(cell.value).strip().replace('\xa0', '')
                        
                        # Try to match each pattern
                        for pattern, key in self.TEXT_PLACEHOLDER_PATTERNS.items():
                            # Check if the cell text matches the pattern
                            match = re.search(pattern, cell_text, re.IGNORECASE)
                            if match:
                                # Found a match - replace with value
                                if key in values and values[key] is not None:
                                    # Replace just the placeholder, keep any other text
                                    new_text = re.sub(pattern, values[key], cell_text, flags=re.IGNORECASE)
                                    cell.value = new_text
                                break  # Move to next cell after first match

    # NEW: Alternative method for direct placeholder replacement
    def replace_direct(self, placeholder_key, value):
        """
        Replace a specific placeholder with a value.
        Useful for dynamic placeholder detection.
        
        Args:
            placeholder_key: The placeholder key (e.g., 'site_id')
            value: The value to insert
        """
        # Convert key to possible placeholder patterns
        patterns = [
            f'\\[{placeholder_key.upper()}\\]',
            f'\\[{placeholder_key.upper().replace("_", "")}\\]',
        ]
        
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell_text = str(cell.value).strip().replace('\xa0', '')
                        
                        for pattern in patterns:
                            if re.search(pattern, cell_text, re.IGNORECASE):
                                new_text = re.sub(pattern, value, cell_text, flags=re.IGNORECASE)
                                cell.value = new_text
                                break