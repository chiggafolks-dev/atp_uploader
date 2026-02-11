# atp_photo_insert.py
import openpyxl
from openpyxl.drawing.image import Image
import os
import re

# Photo placeholder mappings
PHOTO_PLACEHOLDERS = {
    "[PHOTO_FRONT_VIEW]": "Front View",
    "[PHOTO_REAR_VIEW]": "Rear View",
    "[PHOTO_FRONT_SPACE]": "Front Space",
    "[PHOTO_REAR_SPACE]": "Rear Space",
    "[PHOTO_POWER_CABLE]": "Power Cable",
    "[PHOTO_GROUNDING]": "Grounding Cable",
    "[PHOTO_CONNECTION]": "Connection Router",
}

# NEW: Text placeholder patterns - we'll use regex to detect various patterns
TEXT_PLACEHOLDER_PATTERNS = [
    r"\[SITE_?ID\]",  # [SITE_ID] or [SITEID]
    r"\[SITE_?NAME\]",  # [SITE_NAME] or [SITENAME]
    r"\[HOSTNAME\]",
    r"\[SCOPE_?OF_?WORK\]",  # [SCOPE_OF_WORK] or [SCOPEOFWORK]
    r"\[DEVICE_?TYPE\]",  # [DEVICE_TYPE] or [DEVICETYPE]
    r"\[PROJECT_?CODE\]",
    r"\[DATE\]",
    r"\[ENGINEER\]",
    r"\[LOCATION\]",
    r"\[ADDRESS\]",
    r"\[CITY\]",
    r"\[STATE\]",
    r"\[ZIP\]",
    r"\[COUNTRY\]",
    # Add more patterns as needed
]

# NEW: Mapping for clean display names
TEXT_PLACEHOLDER_DISPLAY_NAMES = {
    "[SK_1]" : "Systemkey 1",
    "[SITE_ID1]" : "SITE ID 1",
    "[SITE_NAME1]" : "SITE NAME 1",
    "[SITE_ID]": "Site ID",
    "[SITEID]": "Site ID",
    "[SITE_NAME]": "Site Name",
    "[SITENAME]": "Site Name",
    "[HOSTNAME]": "Hostname",
    "[SCOPE_OF_WORK]": "Scope of Work",
    "[SCOPEOFWORK]": "Scope of Work",
    "[DEVICE_TYPE]": "Device Type",
    "[DEVICETYPE]": "Device Type",
    "[PROJECT_CODE]": "Project Code",
    "[PROJECTCODE]": "Project Code",
    "[DATE]": "Date",
    "[ENGINEER]": "Engineer",
    "[LOCATION]": "Location",
    "[ADDRESS]": "Address",
    "[CITY]": "City",
    "[STATE]": "State",
    "[ZIP]": "ZIP Code",
    "[COUNTRY]": "Country",
}


class ATPPhotoInserter:
    def __init__(self, excel_path):
        self.wb = openpyxl.load_workbook(excel_path)
        self.photo_mappings = self.detect_photo_placeholders()
        # NEW: Detect text placeholders
        self.text_mappings = self.detect_text_placeholders()

    def detect_photo_placeholders(self):
        """Detect photo placeholder cells in the Excel template."""
        mappings = []

        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        text = str(cell.value).strip().replace("\xa0", "")

                        if text in PHOTO_PLACEHOLDERS:
                            mappings.append(
                                {
                                    "sheet": sheet_name,
                                    "photo_type": PHOTO_PLACEHOLDERS[text],
                                    "placeholder": text,
                                    "photo_cell": cell.coordinate,
                                }
                            )

        return mappings

    # NEW: Method to detect text placeholders
    def detect_text_placeholders(self):
        """
        Detect text placeholder cells in the Excel template.
        Uses regex patterns to find various placeholder formats.
        """
        mappings = []

        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        cell_text = str(cell.value).strip().replace("\xa0", "")

                        # Check if the cell text matches any text placeholder pattern
                        for pattern in TEXT_PLACEHOLDER_PATTERNS:
                            match = re.search(pattern, cell_text, re.IGNORECASE)
                            if match:
                                placeholder_text = match.group(
                                    0
                                ).upper()  # Get the matched text

                                # Get a clean display name
                                display_name = TEXT_PLACEHOLDER_DISPLAY_NAMES.get(
                                    placeholder_text,
                                    placeholder_text.strip("[]")
                                    .replace("_", " ")
                                    .title(),
                                )

                                # Generate a key for the placeholder (used in form submission)
                                # Convert [SITE_ID] to 'site_id'
                                placeholder_key = placeholder_text.strip("[]").lower()

                                mappings.append(
                                    {
                                        "sheet": sheet_name,
                                        "placeholder": placeholder_text,
                                        "display_name": display_name,
                                        "placeholder_key": placeholder_key,
                                        "target_cell": cell.coordinate,
                                        "current_value": cell.value,
                                        "description": f"Found in {sheet_name}, cell {cell.coordinate}",
                                    }
                                )
                                # Break after first match to avoid duplicates
                                break

        return mappings

    def get_available_photo_slots(self):
        """Get photo slots information for frontend display."""
        slots = []

        for mapping in self.photo_mappings:
            slots.append(
                {
                    "sheet": mapping["sheet"],
                    "type": mapping["photo_type"],
                    "target_cell": mapping["photo_cell"],
                    "status": "empty",
                }
            )

        return slots

    # NEW: Method to get text fields information
    def get_available_text_fields(self):
        """
        Get text placeholder information for frontend display.

        Returns:
            List of dictionaries with text field information
        """
        text_fields = []

        # Remove duplicates - sometimes same placeholder appears in multiple cells
        seen_placeholders = set()

        for mapping in self.text_mappings:
            placeholder_key = mapping["placeholder_key"]

            # Skip if we've already seen this placeholder
            if placeholder_key in seen_placeholders:
                continue

            seen_placeholders.add(placeholder_key)

            text_fields.append(
                {
                    "placeholder": mapping["placeholder"],
                    "placeholder_key": placeholder_key,
                    "display_name": mapping["display_name"],
                    "sheet": mapping["sheet"],
                    "target_cell": mapping["target_cell"],
                    "required": self.is_field_required(placeholder_key),
                    "description": mapping["description"],
                }
            )

        return text_fields

    # NEW: Helper method to determine if a field is required
    def is_field_required(self, placeholder_key):
        """
        Determine if a field should be marked as required.
        You can customize this logic based on your needs.
        """
        required_fields = {
            "site_id",
            "siteid",
            "site_name",
            "sitename",
            "project_code",
            "projectcode",
        }

        return placeholder_key in required_fields

    # Rest of the existing methods remain the same...
    def insert_photo_by_placeholder(
        self, photo_type, photo_path, resize_width=300, resize_height=200
    ):
        for mapping in self.photo_mappings:
            if mapping["photo_type"] == photo_type:
                ws = self.wb[mapping["sheet"]]

                img = Image(photo_path)
                img.width = resize_width
                img.height = resize_height
                img.anchor = mapping["photo_cell"]

                ws.add_image(img)
                return True

        return False

    def insert_photo_by_cell(
        self, sheet_name, cell, photo_path, resize_width=300, resize_height=200
    ):
        if sheet_name not in self.wb.sheetnames:
            return False

        ws = self.wb[sheet_name]

        img = Image(photo_path)
        img.width = resize_width
        img.height = resize_height
        img.anchor = cell

        ws.add_image(img)
        return True

    def save(self, output_path):
        self.wb.save(output_path)
